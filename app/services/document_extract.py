"""
Document text extraction — supports TXT, MD, PDF, DOCX, XLSX, CSV.

R2 — Table-aware extraction:
  - DOCX: interleaves paragraphs and tables in document order with
    [TABLE N] markers in markdown pipe-delimited format.
  - XLSX: reads workbook sheets via openpyxl, outputs [SHEET: name]
    markers with markdown tables. Skips empty sheets.
  - CSV: stdlib csv with encoding detection (utf-8-sig → utf-8 → latin-1),
    outputs markdown table format.

All paths respect configurable row/col limits
(RAG_EXTRACT_MAX_TABLE_ROWS / RAG_EXTRACT_MAX_TABLE_COLS).

extract_text() signature is unchanged for backward compatibility.
extract_text_with_metadata() returns (text, extraction_metadata_dict).
"""
from __future__ import annotations

import csv
import io
import logging
from dataclasses import dataclass
from io import BytesIO
from typing import Any, Final
import re


logger = logging.getLogger(__name__)


class UnsupportedFileType(Exception):
    """Raised when input format is not supported or required dependency missing."""


class FileTooLarge(Exception):
    """Raised when input file exceeds allowed size."""


class ExtractedTextTooLarge(Exception):
    """Raised when extracted text exceeds allowed size."""


@dataclass(frozen=True)
class ExtractLimits:
    max_bytes: int = 10 * 1024 * 1024        # input file bytes
    max_chars: int = 1_000_000               # rough guard
    max_text_bytes: int = 5 * 1024 * 1024    # OUTPUT utf-8 bytes (MUST match schema)


@dataclass(frozen=True)
class TableLimits:
    """Limits for table/sheet extraction to prevent OOM on huge files."""
    max_rows: int = 500
    max_cols: int = 50
    include_empty_cells: bool = False


_ZERO_WIDTH_RE: Final = re.compile(r"[\u200B-\u200D\uFEFF]")
_NULL_RE: Final = re.compile(r"\x00+")
_WS_RE: Final = re.compile(r"[ \t]+\n")


def _normalize_content_type(content_type: str | None) -> str:
    if not content_type:
        return ""
    return content_type.split(";", 1)[0].strip().lower()


def _decode_text(data: bytes) -> str:
    try:
        return data.decode("utf-8-sig", errors="replace")
    except Exception:
        return data.decode("utf-8", errors="replace")


def _cleanup_text(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = _NULL_RE.sub("", text)
    text = _ZERO_WIDTH_RE.sub("", text)
    text = _WS_RE.sub("\n", text)
    return text.strip()


def _enforce_char_limit(text: str, max_chars: int) -> str:
    if len(text) > max_chars:
        raise ExtractedTextTooLarge(f"Text extracted too large: {len(text)} chars > {max_chars}")
    return text


def _clamp_utf8_bytes(text: str, max_bytes: int) -> str:
    b = text.encode("utf-8", errors="ignore")
    if len(b) <= max_bytes:
        return text

    cut = b[:max_bytes]
    # ensure valid UTF-8 boundary
    while True:
        try:
            return cut.decode("utf-8", errors="strict")
        except UnicodeDecodeError:
            if not cut:
                return ""
            cut = cut[:-1]


def _enforce_text_bytes(text: str, max_text_bytes: int) -> str:
    # clamp (không raise) để router/service không dính 422 schema
    return _clamp_utf8_bytes(text, max_text_bytes)


def _cell_to_str(value: Any) -> str:
    """Convert a cell value to string, handling None, numbers, datetimes."""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        # Avoid trailing .0 for integers stored as float
        if isinstance(value, float) and value == int(value):
            return str(int(value))
        return str(value)
    return str(value).strip()


def _format_table_rows(
    rows: list[list[str]],
    *,
    max_rows: int,
    max_cols: int,
    include_empty_cells: bool,
) -> str:
    """
    Format rows into a markdown pipe-delimited table string.

    Respects max_rows and max_cols. If include_empty_cells is False,
    empty-string cells become a single space for readability.
    """
    if not rows:
        return ""

    # Determine effective column count
    effective_cols = min(
        max(len(row) for row in rows) if rows else 0,
        max_cols,
    )
    if effective_cols == 0:
        return ""

    lines: list[str] = []
    truncated_rows = False

    for row_idx, row in enumerate(rows):
        if row_idx >= max_rows:
            truncated_rows = True
            break

        cells: list[str] = []
        for col_idx in range(effective_cols):
            val = row[col_idx] if col_idx < len(row) else ""
            if not val and not include_empty_cells:
                val = " "
            # Escape pipe chars in cell content
            val = val.replace("|", "\\|")
            cells.append(val)

        lines.append("| " + " | ".join(cells) + " |")

    if truncated_rows:
        lines.append(f"| ... (truncated at {max_rows} rows) |")

    return "\n".join(lines)


# ── DOCX extraction ──────────────────────────────────────────────────


def _extract_docx(
    data: bytes,
    *,
    table_limits: TableLimits,
) -> tuple[str, dict[str, Any]]:
    """
    Extract text from DOCX with interleaved paragraphs and tables.

    Walks the XML body in document order so paragraphs and tables
    appear in their original sequence.

    Returns (text, extraction_metadata).
    """
    try:
        from docx import Document
        from docx.table import Table
        from docx.text.paragraph import Paragraph
        from docx.oxml.ns import qn
    except Exception as e:
        raise UnsupportedFileType(
            "DOCX chưa hỗ trợ (thiếu python-docx). Cài: pip install python-docx"
        ) from e

    d = Document(BytesIO(data))
    parts: list[str] = []
    table_count = 0

    # Walk body children in document order to interleave paragraphs and tables
    body = d.element.body
    for child in body:
        tag = child.tag

        if tag == qn("w:p"):
            # Paragraph element
            para = Paragraph(child, body)
            text = (para.text or "").strip()
            if text:
                parts.append(text)

        elif tag == qn("w:tbl"):
            # Table element
            table_count += 1
            table = Table(child, body)
            rows_data: list[list[str]] = []

            for row_idx, row in enumerate(table.rows):
                if row_idx >= table_limits.max_rows:
                    break
                cells = [
                    _cell_to_str(cell.text)
                    for cell in row.cells[:table_limits.max_cols]
                ]
                rows_data.append(cells)

            table_text = _format_table_rows(
                rows_data,
                max_rows=table_limits.max_rows,
                max_cols=table_limits.max_cols,
                include_empty_cells=table_limits.include_empty_cells,
            )
            if table_text:
                parts.append(f"[TABLE {table_count}]\n{table_text}")

    meta = {
        "extractor_version": "table_aware_v1",
        "tables_count": table_count,
        "sheets_count": 0,
        "has_tables": table_count > 0,
    }

    return "\n\n".join(parts), meta


# ── XLSX extraction ──────────────────────────────────────────────────


def _extract_xlsx(
    data: bytes,
    *,
    table_limits: TableLimits,
) -> tuple[str, dict[str, Any]]:
    """
    Extract text from XLSX workbook. Each non-empty sheet becomes a
    [SHEET: name] section with markdown table rows.

    Returns (text, extraction_metadata).
    """
    try:
        import openpyxl
    except Exception as e:
        raise UnsupportedFileType(
            "XLSX chưa hỗ trợ (thiếu openpyxl). Cài: pip install openpyxl"
        ) from e

    try:
        wb = openpyxl.load_workbook(
            BytesIO(data),
            read_only=True,
            data_only=True,
        )
    except Exception as e:
        raise UnsupportedFileType(
            f"Không thể đọc file XLSX: {type(e).__name__}"
        ) from e

    parts: list[str] = []
    sheets_count = 0
    tables_count = 0

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Detect data bounds
            if ws.max_row is None or ws.max_row == 0:
                continue

            rows_data: list[list[str]] = []
            row_count = 0

            for row in ws.iter_rows(
                min_row=ws.min_row,
                max_row=min(ws.max_row, ws.min_row + table_limits.max_rows - 1),
                max_col=min(ws.max_column or 1, table_limits.max_cols),
                values_only=True,
            ):
                # Skip entirely empty rows
                if all(v is None for v in row):
                    continue

                cells = [_cell_to_str(v) for v in row]
                rows_data.append(cells)
                row_count += 1

                if row_count >= table_limits.max_rows:
                    break

            if not rows_data:
                continue

            sheets_count += 1
            tables_count += 1

            table_text = _format_table_rows(
                rows_data,
                max_rows=table_limits.max_rows,
                max_cols=table_limits.max_cols,
                include_empty_cells=table_limits.include_empty_cells,
            )
            if table_text:
                parts.append(f"[SHEET: {sheet_name}]\n{table_text}")

    finally:
        wb.close()

    meta = {
        "extractor_version": "table_aware_v1",
        "tables_count": tables_count,
        "sheets_count": sheets_count,
        "has_tables": tables_count > 0,
    }

    return "\n\n".join(parts), meta


# ── CSV extraction ───────────────────────────────────────────────────


def _detect_csv_encoding(data: bytes) -> str:
    """Try decoding with common Vietnamese/international encodings."""
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            data.decode(enc)
            return enc
        except (UnicodeDecodeError, ValueError):
            continue
    return "latin-1"  # fallback — always succeeds


def _extract_csv(
    data: bytes,
    *,
    table_limits: TableLimits,
) -> tuple[str, dict[str, Any]]:
    """
    Extract text from CSV file. Outputs markdown table format.

    Returns (text, extraction_metadata).
    """
    encoding = _detect_csv_encoding(data)

    try:
        text_content = data.decode(encoding, errors="replace")
    except Exception:
        text_content = data.decode("latin-1", errors="replace")

    # Detect dialect
    try:
        sample = text_content[:8192]
        dialect = csv.Sniffer().sniff(sample)
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text_content), dialect=dialect)

    rows_data: list[list[str]] = []
    row_count = 0

    for row in reader:
        if row_count >= table_limits.max_rows:
            break

        # Limit columns
        cells = [
            _cell_to_str(v) for v in row[:table_limits.max_cols]
        ]

        # Skip entirely empty rows
        if all(not c.strip() for c in cells):
            continue

        rows_data.append(cells)
        row_count += 1

    table_text = _format_table_rows(
        rows_data,
        max_rows=table_limits.max_rows,
        max_cols=table_limits.max_cols,
        include_empty_cells=table_limits.include_empty_cells,
    )

    tables_count = 1 if rows_data else 0

    meta = {
        "extractor_version": "table_aware_v1",
        "tables_count": tables_count,
        "sheets_count": 0,
        "has_tables": tables_count > 0,
    }

    return table_text, meta


# ── Public API ────────────────────────────────────────────────────────


_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
_CSV_MIMES = {"text/csv", "application/csv"}


def extract_text_with_metadata(
    filename: str | None,
    content_type: str | None,
    data: bytes,
    *,
    limits: ExtractLimits = ExtractLimits(),
    table_limits: TableLimits = TableLimits(),
) -> tuple[str, dict[str, Any]]:
    """
    Extract text and return (text, extraction_metadata).

    R2 extension of extract_text() — same logic but also returns
    structured metadata about the extraction (table counts, etc.).
    """
    if len(data) > limits.max_bytes:
        raise FileTooLarge(f"File too large: {len(data)} bytes > {limits.max_bytes}")

    name = (filename or "").lower().strip()
    ctype = _normalize_content_type(content_type)

    # ── Extraction metadata defaults ──
    default_meta: dict[str, Any] = {
        "extractor_version": "table_aware_v1",
        "tables_count": 0,
        "sheets_count": 0,
        "has_tables": False,
    }

    # TXT / MD
    if name.endswith(".txt") or name.endswith(".md") or (ctype in {"text/plain", "text/markdown"}):
        text = _cleanup_text(_decode_text(data))
        text = _enforce_char_limit(text, limits.max_chars)
        return _enforce_text_bytes(text, limits.max_text_bytes), default_meta

    # PDF
    if name.endswith(".pdf") or (ctype == "application/pdf"):
        try:
            import fitz  # PyMuPDF
        except Exception as e:
            raise UnsupportedFileType("PDF chưa hỗ trợ (thiếu pymupdf). Cài: pip install pymupdf") from e

        pdf_parts: list[str] = []
        doc = fitz.open(stream=data, filetype="pdf")
        try:
            for page in doc:
                pdf_parts.append(page.get_text("text") or "")
        finally:
            doc.close()

        text = _cleanup_text("\n".join(pdf_parts))
        text = _enforce_char_limit(text, limits.max_chars)
        return _enforce_text_bytes(text, limits.max_text_bytes), default_meta

    # DOCX — table-aware (R2)
    if name.endswith(".docx") or (ctype == _DOCX_MIME):
        text, extraction_meta = _extract_docx(data, table_limits=table_limits)
        text = _cleanup_text(text)
        text = _enforce_char_limit(text, limits.max_chars)
        return _enforce_text_bytes(text, limits.max_text_bytes), extraction_meta

    # XLSX — new in R2
    if name.endswith(".xlsx") or (ctype == _XLSX_MIME):
        text, extraction_meta = _extract_xlsx(data, table_limits=table_limits)
        text = _cleanup_text(text)
        text = _enforce_char_limit(text, limits.max_chars)
        return _enforce_text_bytes(text, limits.max_text_bytes), extraction_meta

    # CSV — new in R2
    if name.endswith(".csv") or (ctype in _CSV_MIMES):
        text, extraction_meta = _extract_csv(data, table_limits=table_limits)
        text = _cleanup_text(text)
        text = _enforce_char_limit(text, limits.max_chars)
        return _enforce_text_bytes(text, limits.max_text_bytes), extraction_meta

    # DOC legacy
    if name.endswith(".doc") or (ctype == "application/msword"):
        raise UnsupportedFileType("File .doc (Word cũ) chưa hỗ trợ. Hãy chuyển sang .docx hoặc bật convert bằng LibreOffice.")

    raise UnsupportedFileType(
        f"Định dạng không hỗ trợ. filename={filename!r}, content_type={content_type!r}. "
        f"Hỗ trợ: txt, md, pdf, docx, xlsx, csv."
    )


def extract_text(
    filename: str | None,
    content_type: str | None,
    data: bytes,
    *,
    limits: ExtractLimits = ExtractLimits(),
) -> str:
    """
    Extract text from document. Backward-compatible R1 API.

    Returns extracted text as a single string.
    Delegates to extract_text_with_metadata() internally.
    """
    text, _meta = extract_text_with_metadata(
        filename,
        content_type,
        data,
        limits=limits,
    )
    return text